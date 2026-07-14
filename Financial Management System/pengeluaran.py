import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import datetime
import os

FILE_NAME = "pengeluaran.xlsx"
selected_item = None

# Simpan ke Excel

def simpan_pengeluaran(tanggal, kategori, jumlah, keterangan):
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pengeluaran"
        ws.append(["Tanggal", "Kategori", "Jumlah", "Keterangan"])
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([tanggal, kategori, jumlah, keterangan])
    wb.save(FILE_NAME)


# Load Excel

def load_pengeluaran():
    data = []

    if not os.path.exists(FILE_NAME):
        return data

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))

    return data


# Simpan ulang Excel

def save_all(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pengeluaran"

    ws.append(["Tanggal", "Kategori", "Jumlah", "Keterangan"])

    for row in data:
        ws.append(row)

    wb.save(FILE_NAME)


# Tambah Data

def tambah_pengeluaran():
    tanggal = datetime.date.today().strftime("%Y-%m-%d")
    kategori = entry_kategori.get()
    jumlah = entry_jumlah.get()
    keterangan = entry_keterangan.get()

    if kategori == "" or jumlah == "":
        messagebox.showwarning("Peringatan", "Kategori dan jumlah harus diisi!")
        return

    try:
        jumlah = float(jumlah)
    except:
        messagebox.showwarning("Peringatan", "Jumlah harus berupa angka!")
        return

    simpan_pengeluaran(tanggal, kategori, jumlah, keterangan)

    tree.insert(
        "",
        "end",
        values=(tanggal, kategori, f"Rp {jumlah:,.0f}", keterangan),
    )

    entry_kategori.delete(0, tk.END)
    entry_jumlah.delete(0, tk.END)
    entry_keterangan.delete(0, tk.END)

    update_total()

    messagebox.showinfo("Sukses", "Data berhasil ditambahkan!")


# =========================
# PILIH DATA
# =========================
def pilih_data(event):
    global selected_item

    selected = tree.selection()
    if not selected:
        return

    selected_item = selected[0]
    values = tree.item(selected_item)["values"]

    entry_kategori.delete(0, tk.END)
    entry_kategori.insert(0, values[1])

    entry_jumlah.delete(0, tk.END)
    entry_jumlah.insert(0, str(values[2]).replace("Rp", "").replace(",", "").strip())

    entry_keterangan.delete(0, tk.END)
    entry_keterangan.insert(0, values[3])


# =========================
# EDIT DATA
# =========================
def edit_pengeluaran():
    global selected_item

    if selected_item is None:
        messagebox.showwarning("Peringatan", "Pilih data yang ingin diedit!")
        return

    tanggal = tree.item(selected_item)["values"][0]
    kategori = entry_kategori.get()
    keterangan = entry_keterangan.get()

    try:
        jumlah = float(entry_jumlah.get())
    except:
        messagebox.showwarning("Peringatan", "Jumlah harus berupa angka!")
        return

    tree.item(
        selected_item,
        values=(tanggal, kategori, f"Rp {jumlah:,.0f}", keterangan)
    )

    data = []

    for item in tree.get_children():
        row = tree.item(item)["values"]
        data.append([
            row[0],
            row[1],
            float(str(row[2]).replace("Rp", "").replace(",", "").strip()),
            row[3]
        ])

    save_all(data)

    entry_kategori.delete(0, tk.END)
    entry_jumlah.delete(0, tk.END)
    entry_keterangan.delete(0, tk.END)

    selected_item = None

    update_total()

    messagebox.showinfo("Sukses", "Data berhasil diedit!")

# Hapus Data

def hapus_pengeluaran():
    selected = tree.selection()

    if not selected:
        messagebox.showwarning("Peringatan", "Pilih data yang ingin dihapus!")
        return

    if not messagebox.askyesno("Konfirmasi", "Yakin ingin menghapus data?"):
        return

    values = tree.item(selected)["values"]

    tree.delete(selected)

    data = load_pengeluaran()

    data_baru = []

    for row in data:
        if not (
            row[0] == values[0]
            and row[1] == values[1]
            and float(row[2]) == float(str(values[2]).replace("Rp", "").replace(",", "").strip())
            and row[3] == values[3]
        ):
            data_baru.append(row)

    save_all(data_baru)

    update_total()

    messagebox.showinfo("Sukses", "Data berhasil dihapus!")


# Hitung Total

def update_total():
    total = 0

    for row in load_pengeluaran():
        total += float(row[2])

    label_total.config(
        text=f"💰 Total Pengeluaran : Rp {total:,.0f}"
    )
    


# GUI

root = tk.Tk()
root.title("🌸 Aplikasi Pengeluaran Uang 🌸")
root.geometry("1200x700")
root.resizable(False, False)
root.configure(bg="#fceef5")

frame_input = tk.Frame(root, bg="#fceef5", padx=10, pady=10)
frame_input.pack(fill="x")

# Atur kolom agar form berada di tengah
frame_input.columnconfigure(0, weight=1)
frame_input.columnconfigure(1, weight=0)
frame_input.columnconfigure(2, weight=0)
frame_input.columnconfigure(3, weight=1)

tk.Label(frame_input, text="Kategori :", bg="#fceef5",
         font=("Comic Sans MS",12,"bold")).grid(
             row=0, column=0, sticky="w", padx=20, pady=8)

entry_kategori = tk.Entry(frame_input,
                          font=("Comic Sans MS",12),
                          width=45)
entry_kategori.grid(row=0, column=1, sticky="w", padx=10, pady=8)


tk.Label(frame_input, text="Jumlah (Rp) :", bg="#fceef5",
         font=("Comic Sans MS",12,"bold")).grid(
             row=1, column=0, sticky="w", padx=20, pady=8)

entry_jumlah = tk.Entry(frame_input,
                        font=("Comic Sans MS",12),
                        width=45)
entry_jumlah.grid(row=1, column=1, sticky="w", padx=10, pady=8)


tk.Label(frame_input, text="Keterangan :", bg="#fceef5",
         font=("Comic Sans MS",12,"bold")).grid(
             row=2, column=0, sticky="w", padx=20, pady=8)

entry_keterangan = tk.Entry(frame_input,
                            font=("Comic Sans MS",12),
                            width=45)
entry_keterangan.grid(row=2, column=1, sticky="w", padx=10, pady=8)





btn_tambah = tk.Button(
    frame_input,
    text="➕ Tambah",
    command=tambah_pengeluaran,
    bg="#74b9ff",
    fg="white",
    font=("Comic Sans MS",11,"bold"),
    width=13
)
btn_tambah.grid(row=3,column=0,sticky="w",padx=20,pady=20)

btn_edit = tk.Button(
    frame_input,
    text="✏️ Edit",
    command=edit_pengeluaran,
    bg="#f1c40f",
    fg="white",
    font=("Comic Sans MS",11,"bold"),
    width=13
)
btn_edit.grid(row=3,column=1,pady=20)

btn_hapus = tk.Button(
    frame_input,
    text="🗑️ Hapus",
    command=hapus_pengeluaran,
    bg="#ff7675",
    fg="white",
    font=("Comic Sans MS",11,"bold"),
    width=13
)
btn_hapus.grid(row=3, column=3, sticky="e", padx=20, pady=20)


judul = tk.Label(
    root,
    text="📋 DATA PENGELUARAN",
    bg="#fceef5",
    fg="#d63031",
    font=("Comic Sans MS",16,"bold")
)

judul.pack(pady=(0,10))

frame_tabel = tk.Frame(root)
frame_tabel.pack(fill="both", expand=True)

kolom = ("Tanggal", "Kategori", "Jumlah", "Keterangan")

tree = ttk.Treeview(frame_tabel, columns=kolom, show="headings")

style = ttk.Style()

style.theme_use("clam")

style.configure(
    "Treeview.Heading",
    font=("Comic Sans MS",11,"bold"),
    background="#ffb6c1",
    foreground="black"
)

style.configure(
    "Treeview",
    rowheight=28,
    font=("Comic Sans MS",10)
)

for col in kolom:
    tree.heading(col, text=col)
    tree.column(col, width=210, anchor="center")

tree.pack(fill="both", expand=True)
tree.bind("<<TreeviewSelect>>", pilih_data)

frame_total = tk.Frame(root, bg="#fceef5")
frame_total.pack(fill="x", side="bottom")

label_total = tk.Label(
    frame_total,
    text="💰 Total Pengeluaran : Rp 0",
   font=("Comic Sans MS", 9, "bold"),
    bg="#fceef5",
    fg="#d63031",
    anchor="center"
)

label_total.pack(fill="x", pady=10)

for row in load_pengeluaran():
    tree.insert(
        "",
        "end",
        values=(row[0], row[1], f"Rp {float(row[2]):,.0f}", row[3]),
    )

update_total()

root.mainloop()